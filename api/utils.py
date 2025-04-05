from string import ascii_uppercase
from typing import List

from django.db.models import Manager
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import quote_sheetname
from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from pandas.core.series import Series

from .models import *


def create_example(row_count: int, high_school: HighSchool) -> Workbook:
    wb = Workbook()
    wb.remove(wb.worksheets[0])

    font = Font(name="Times New Roman", size=11, color="000000")
    alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True, wrapText=True
    )
    bold_font = Font(name="Times New Roman", size=11, color="000000", bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws: Worksheet = wb.create_sheet("Maglumat")

    column_headers = [
        "T/B",
        "F.A.Aa",
        "Doglan senesi",
        "Jynsy",
        "Welaýaty",
        "Milleti",
        "Ýurdy",
        "Hünari",
        "Kursy",
        "Töleg görnüşi",
        "Maşgala ýagdaýy",
        "Ýazgyda duran salgysy",
        "Öý, iş, mobil telefony",
        "Pasport seriýasy, belgisi",
        "Harby borç",
        "Okuwa giren senesi",
        "Bellikler",
    ]
    column_index = 0

    for column_id in ascii_uppercase[:17]:
        ws[f"{column_id}1"].value = column_headers[column_index]
        column_index += 1
        ws.column_dimensions[column_id].width = 20
        for row_id in range(1, row_count + 2):
            ws[f"{column_id}{row_id}"].border = border
            ws[f"{column_id}{row_id}"].alignment = alignment
            ws[f"{column_id}{row_id}"].font = bold_font if row_id == 1 else font
            ws.row_dimensions[row_id].height = 50

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 20

    # Specialization validation values
    index = 1
    specialization_ws: Worksheet = wb.create_sheet("Hünarler")
    for specialization in Specialization.objects.filter(
        id__in=[
            department_specialization.specialization.id
            for department_specialization in DepartmentSpecialization.objects.filter(
                faculty_department__high_school_faculty__high_school=high_school
            )
        ]
    ):
        specialization_ws[f"A{index}"].value = specialization.name
        index += 1

    # Country validation values
    index = 1
    country_ws: Worksheet = wb.create_sheet("Ýurtlar")
    for country in Country.objects.all():
        country_ws[f"A{index}"].value = country.name
        index += 1

    # Nationality validation values
    index = 1
    nationality_ws: Worksheet = wb.create_sheet("Milletler")
    for nationality in Nationality.objects.all():
        nationality_ws[f"A{index}"].value = nationality.name
        index += 1

    # Region validation values
    index = 1
    region_ws: Worksheet = wb.create_sheet("Welaýatlar")
    for region in Region.objects.all():
        region_ws[f"A{index}"].value = region.name
        index += 1

    # Gender validation values
    gender_ws: Worksheet = wb.create_sheet("Jynslar")
    gender_ws["A1"].value = "Oglan"
    gender_ws["A2"].value = "Gyz"

    # Course validation values
    study_year_ws: Worksheet = wb.create_sheet("Kurslar")
    study_year_ws["A1"].value = "DÖB"
    for i in range(7):
        study_year_ws[f"A{i + 2}"].value = str(i + 1)

    # Family status validation values
    family_status_ws: Worksheet = wb.create_sheet("Maşgala ýagdaýlary")
    family_status_ws["A1"].value = "Hossarly"
    family_status_ws["A2"].value = "Ýarym ýetim"
    family_status_ws["A3"].value = "Doly ýetim"
    family_status_ws["A4"].value = "Ýetimler öýünde ösen"

    # Payment type validation values
    payment_type_ws: Worksheet = wb.create_sheet("Töleg görnüşleri")
    payment_type_ws["A1"].value = "Tölegli"
    payment_type_ws["A2"].value = "Býudjet"

    specialization_data_val = DataValidation(
        type="list",
        formula1=f"={quote_sheetname('Hünarler')}!$A:$A",
    )
    country_data_val = DataValidation(
        type="list",
        formula1=f"={quote_sheetname('Ýurtlar')}!$A:$A",
    )
    nationality_data_val = DataValidation(
        type="list",
        formula1=f"={quote_sheetname('Milletler')}!$A:$A",
    )
    region_data_val = DataValidation(
        type="list",
        formula1=f"={quote_sheetname('Welaýatlar')}!$A:$A",
    )
    gender_data_val = DataValidation(
        type="list",
        formula1=f"={quote_sheetname('Jynslar')}!$A:$A",
    )
    study_year_data_val = DataValidation(
        type="list",
        formula1=f"={quote_sheetname('Kurslar')}!$A:$A",
    )
    family_status_data_val = DataValidation(
        type="list",
        formula1=f"={quote_sheetname('Maşgala ýagdaýlary')}!$A:$A",
    )
    payment_type_data_val = DataValidation(
        type="list",
        formula1=f"={quote_sheetname('Töleg görnüşleri')}!$A:$A",
    )

    ws.add_data_validation(specialization_data_val)
    ws.add_data_validation(country_data_val)
    ws.add_data_validation(nationality_data_val)
    ws.add_data_validation(region_data_val)
    ws.add_data_validation(gender_data_val)
    ws.add_data_validation(study_year_data_val)
    ws.add_data_validation(family_status_data_val)
    ws.add_data_validation(payment_type_data_val)
    for row_index in range(2, row_count + 2):
        ws[f"A{row_index}"].value = row_index - 1
        specialization_data_val.add(ws[f"H{row_index}"])
        country_data_val.add(ws[f"G{row_index}"])
        region_data_val.add(ws[f"E{row_index}"])
        gender_data_val.add(ws[f"D{row_index}"])
        study_year_data_val.add(ws[f"I{row_index}"])
        family_status_data_val.add(ws[f"K{row_index}"])
        nationality_data_val.add(ws[f"F{row_index}"])
        payment_type_data_val.add(ws[f"J{row_index}"])

    return wb


def validate_not_null_field(row: Series, excepted_fields: List[str]):
    not_valid_fields = []

    for key in row.keys():
        if row.isnull()[key]:
            if not key in excepted_fields:
                not_valid_fields.append(key)

    return (False, not_valid_fields) if len(not_valid_fields) > 0 else (True, [])


def advanced_quantity_filter(payload: dict):
    print(payload)
    students = Student.objects.filter(active=True, is_expelled=False)
    query_set = Student.objects.none()
    if len(payload["high_schools"]):
        for high_school_id in payload["high_schools"]:
            query_set = query_set.union(students.filter(high_school__id=high_school_id))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["faculties"]):
        for faculty_id in payload["faculties"]:
            query_set = query_set.union(
                students.filter(
                    specialization__faculty_department__high_school_faculty__faculty__id=faculty_id
                )
            )
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["departments"]):
        for department_id in payload["departments"]:
            query_set = query_set.union(
                students.filter(
                    specialization__faculty_department__department__id=department_id
                )
            )
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["specializations"]):
        for specialization_id in payload["specializations"]:
            query_set = query_set.union(
                students.filter(specialization__specialization__id=specialization_id)
            )
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["study_years"]):
        for study_year in payload["study_years"]:
            query_set = query_set.union(students.filter(study_year=study_year))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["payment_types"]):
        for payment_type_id in payload["payment_types"]:
            payment_type = "P" if payment_type_id == 1 else "B"
            query_set = query_set.union(students.filter(payment_type=payment_type))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["genders"]):
        for gender_id in payload["genders"]:
            gender = "M" if gender_id == 1 else "F"
            query_set = query_set.union(students.filter(gender=gender))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["nationalities"]):
        for nationality_id in payload["nationalities"]:
            query_set = query_set.union(students.filter(nationality__id=nationality_id))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["countries"]):
        for country_id in payload["countries"]:
            query_set = query_set.union(students.filter(country__id=country_id))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["regions"]):
        for region_id in payload["regions"]:
            query_set = query_set.union(students.filter(region__id=region_id))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if payload["military_service"] == 1:
        query_set = query_set.union(students.exclude(military_service=None))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    elif payload["military_service"] == 2:
        query_set = query_set.union(students.filter(military_service=None, gender="M"))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()

    print(students)
    degrees = [
        {
            "name": "Bakalawr",
            "count": students.filter(
                specialization__specialization__degree__name__contains="bakalawr"
            ).count(),
        },
        {
            "name": "Hünärmen",
            "count": students.filter(
                specialization__specialization__degree__name__contains="hünärmen"
            ).count(),
        },
        {
            "name": "Aspirantura",
            "count": students.filter(
                specialization__specialization__degree__name__contains="aspirantura"
            ).count(),
        },
        {
            "name": "Magistratura",
            "count": students.filter(
                specialization__specialization__degree__name__contains="magistratura"
            ).count(),
        },
        {
            "name": "Bakalawr tölegli",
            "count": students.filter(
                specialization__specialization__degree__name__contains="bakalawr",
                payment_type="P",
            ).count(),
        },
        {
            "name": "Hünärmen tölegli",
            "count": students.filter(
                specialization__specialization__degree__name__contains="hünärmen",
                payment_type="P",
            ).count(),
        },
        {
            "name": "Aspirantura tölegli",
            "count": students.filter(
                specialization__specialization__degree__name__contains="aspirantura",
                payment_type="P",
            ).count(),
        },
        {
            "name": "Magistratura tölegli",
            "count": students.filter(
                specialization__specialization__degree__name__contains="magistratura",
                payment_type="P",
            ).count(),
        },
        {
            "name": "Bakalawr tölegsiz",
            "count": students.filter(
                specialization__specialization__degree__name__contains="bakalawr",
                payment_type="B",
            ).count(),
        },
        {
            "name": "Hünärmen tölegsiz",
            "count": students.filter(
                specialization__specialization__degree__name__contains="hünärmen",
                payment_type="B",
            ).count(),
        },
        {
            "name": "Aspirantura tölegsiz",
            "count": students.filter(
                specialization__specialization__degree__name__contains="aspirantura",
                payment_type="B",
            ).count(),
        },
        {
            "name": "Magistratura tölegsiz",
            "count": students.filter(
                specialization__specialization__degree__name__contains="magistratura",
                payment_type="B",
            ).count(),
        },
    ]
    return {
        "students": [
            {"name": "Jemi", "query": 0, "count": students.count()},
            {
                "name": "Oglan",
                "query": 1,
                "count": students.filter(gender="M").count(),
            },
            {
                "name": "Gyz",
                "query": 2,
                "count": students.filter(gender="F").count(),
            },
        ],
        "regions": [
            {
                "id": region.id,
                "name": region.name,
                "count": students.filter(region=region).count(),
            }
            for region in Region.objects.all()
        ],
        "degrees": degrees,
    }


def advanced_filter(payload: dict, key: str = "gender", value: int = 0):
    students = Student.objects.filter(active=True, is_expelled=False, is_obsolete=False)
    query_set = Student.objects.none()
    if len(payload["high_schools"]):
        for high_school_id in payload["high_schools"]:
            query_set = query_set.union(students.filter(high_school__id=high_school_id))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["faculties"]):
        for faculty_id in payload["faculties"]:
            query_set = query_set.union(
                students.filter(
                    specialization__faculty_department__high_school_faculty__faculty__id=faculty_id
                )
            )
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["departments"]):
        for department_id in payload["departments"]:
            query_set = query_set.union(
                students.filter(
                    specialization__faculty_department__department__id=department_id
                )
            )
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["specializations"]):
        for specialization_id in payload["specializations"]:
            query_set = query_set.union(
                students.filter(specialization__specialization__id=specialization_id)
            )
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["study_years"]):
        for study_year in payload["study_years"]:
            query_set = query_set.union(students.filter(study_year=study_year))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["payment_types"]):
        for payment_type_id in payload["payment_types"]:
            payment_type = "P" if payment_type_id == 1 else "B"
            query_set = query_set.union(students.filter(payment_type=payment_type))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["genders"]):
        for gender_id in payload["genders"]:
            gender = "M" if gender_id == 1 else "F"
            query_set = query_set.union(students.filter(gender=gender))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["nationalities"]):
        for nationality_id in payload["nationalities"]:
            query_set = query_set.union(students.filter(nationality__id=nationality_id))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["countries"]):
        for country_id in payload["countries"]:
            query_set = query_set.union(students.filter(country__id=country_id))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if len(payload["regions"]):
        for region_id in payload["regions"]:
            query_set = query_set.union(students.filter(region__id=region_id))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    if payload["military_service"] == 1:
        query_set = query_set.union(
            students.exclude(military_service=None).exclude(military_service="")
        )
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()
    elif payload["military_service"] == 2:
        query_set = query_set.union(students.filter(military_service=None, gender="M"))
        students = students.filter(id__in=[student.id for student in query_set.all()])
        query_set = Student.objects.none()

    if (
        not len(payload["high_schools"])
        and not len(payload["faculties"])
        and not len(payload["departments"])
        and not len(payload["specializations"])
        and not len(payload["study_years"])
        and not len(payload["payment_types"])
        and not len(payload["genders"])
        and not len(payload["nationalities"])
        and not len(payload["countries"])
        and not len(payload["regions"])
        and payload["military_service"] == 0
    ):
        students = students.none()

    if key == "gender":
        if value == 1:
            students = students.filter(gender="M")
        elif value == 2:
            students = students.filter(gender="F")
    if key == "region":
        students = students.filter(region=value)
    return [
        {
            "id": student.id,
            "full_name": student.full_name,
            "high_school": student.high_school.name,
        }
        for student in students
    ]


def filter_by_query(data: Manager[Student], query_dict):
    if len(list(query_dict.keys())):
        key = list(query_dict.keys())[0]
        value = int(query_dict[list(query_dict.keys())[0]])
        match key:
            case "faculty":
                data = data.filter(
                    specialization__faculty_department__high_school_faculty__faculty__id=value
                )
            case "department":
                data = data.filter(
                    specialization__faculty_department__department__id=value
                )
            case "specialization":
                data = data.filter(specialization__specialization__id=value)
            case "nationality":
                data = data.filter(nationality__id=value)
            case "country":
                data = data.filter(country__id=value)
            case "region":
                data = data.filter(region__id=value)
            case "degree":
                data = data.filter(specialization__specialization__degree__id=value)
            case "classificator":
                data = data.filter(
                    specialization__specialization__classificator__id=value
                )
    return data


def validate_excel_fields(
    row, index: int, high_school: HighSchool, row_validation: bool, dry_run: bool = True
):
    data = {}
    invalid_fields = []
    data["full_name"] = row["F.A.Aa"]
    if type(row["Doglan senesi"]) == datetime.datetime:
        data["birth_date"] = row["Doglan senesi"]
    elif type(row["Doglan senesi"]) == str:
        try:
            data["birth_date"] = datetime.datetime.strptime(
                row["Doglan senesi"].rstrip().lstrip(), "%d.%M.%Y"
            )
        except:
            print(f'"{row["Doglan senesi"]}"')
            invalid_fields.append(
                f"Setir №{index + 1}: 'Doglan senesi' meýdançasynda ýalňyşlyk goýberildi"
            )
            row_validation = False
    else:
        data["birth_date"] = row["Doglan senesi"].to_pydatetime()

    if row["Jynsy"].lower() in ("oglan", "gyz"):
        data["gender"] = "M" if row["Jynsy"].lower() == "oglan" else "F"
    else:
        invalid_fields.append(
            f"Setir №{index + 1}: 'Jynsy' meýdançasynda ýalňyşlyk goýberildi"
        )
        row_validation = False

    if Region.objects.filter(name=row["Welaýaty"]).exists():
        data["region"] = Region.objects.get(name=row["Welaýaty"])
    else:
        print(row["Welaýaty"])
        invalid_fields.append(
            f"Setir №{index + 1}: 'Welaýaty' meýdançasynda ýalňyşlyk goýberildi"
        )
        row_validation = False

    if Nationality.objects.filter(name=row["Milleti"]).exists():
        data["nationality"] = Nationality.objects.get(name=row["Milleti"])
    else:
        invalid_fields.append(
            f"Setir №{index + 1}: 'Milleti' meýdançasynda ýalňyşlyk goýberildi"
        )
        row_validation = False

    if Country.objects.filter(name=row["Ýurdy"]).exists():
        data["country"] = Country.objects.get(name=row["Ýurdy"])
    else:
        invalid_fields.append(
            f"Setir №{index + 1}: 'Ýurdy' meýdançasynda ýalňyşlyk goýberildi"
        )
        row_validation = False

    if DepartmentSpecialization.objects.filter(
        specialization__name=row["Hünari"],
        faculty_department__high_school_faculty__high_school=high_school,
    ):
        data["specialization"] = DepartmentSpecialization.objects.get(
            specialization__name=row["Hünari"],
            faculty_department__high_school_faculty__high_school=high_school,
        )
    else:
        invalid_fields.append(
            f"Setir №{index + 1}: 'Hünäri' meýdançasynda ýalňyşlyk goýberildi"
        )
        row_validation = False
    if (
        type(row["Kursy"]) == int
        or type(row["Kursy"]) == float
        or type(row["Kursy"]) == str
    ):
        if type(row["Kursy"]) == float:
            data["course"] = str(int(row["Kursy"]))
        else:
            data["course"] = str(row["Kursy"])
    else:
        invalid_fields.append(
            f"Setir №{index + 1}: 'Kursy' meýdançasynda ýalňyşlyk goýberildi"
        )
        row_validation = False

    if row["Töleg görnüşi"].lower() in ("tölegli", "býudjet"):
        data["payment_type"] = "P" if row["Töleg görnüşi"] == "Tölegli" else "B"
    else:
        invalid_fields.append(
            f"Setir №{index + 1}: 'Töleg görnüşi' meýdançasynda ýalňyşlyk goýberildi"
        )
        row_validation = False

    family_statuses = {
        "hossarly": "FR",
        "ýarym ýetim": "HO",
        "doly ýetim": "CO",
        "ýetimler öýünde ösen": "OE",
    }
    if row["Maşgala ýagdaýy"].lower() in family_statuses.keys():
        data["family_status"] = family_statuses[row["Maşgala ýagdaýy"].lower()]
    else:
        invalid_fields.append(
            f"Setir №{index + 1}: 'Maşgala ýagdaýy' meýdançasynda ýalňyşlyk goýberildi"
        )
        row_validation = False

    if type(row["Ýazgyda duran salgysy"]) == str:
        data["registered_place"] = row["Ýazgyda duran salgysy"]
    else:
        invalid_fields.append(
            f"Setir №{index + 1}: 'Ýazgyda duran salgysy' meýdançasynda ýalňyşlyk goýberildi"
        )
        row_validation = False

    if (
        type(row["Öý, iş, mobil telefony"]) == str
        or type(row["Öý, iş, mobil telefony"]) == int
        or type(row["Öý, iş, mobil telefony"]) == float
    ):
        if type(row["Öý, iş, mobil telefony"]) == float:
            data["phone_number"] = str(int(row["Öý, iş, mobil telefony"]))
        else:
            data["phone_number"] = str(row["Öý, iş, mobil telefony"])
    else:
        invalid_fields.append(
            f"Setir №{index + 1}: 'Öý, iş, mobil telefony' meýdançasynda ýalňyşlyk goýberildi"
        )
        row_validation = False

    if type(row["Pasport seriýasy, belgisi"]) == str:
        data["passport"] = row["Pasport seriýasy, belgisi"]
    else:
        invalid_fields.append(
            f"Setir №{index + 1}: 'Pasport seriýasy, belgisi' meýdançasynda ýalňyşlyk goýberildi"
        )
        row_validation = False

    if type(row["Okuwa giren senesi"]) == datetime.datetime:
        data["admission_date"] = row["Okuwa giren senesi"]
    elif type(row["Okuwa giren senesi"]) == str:
        try:
            data["admission_date"] = datetime.datetime.strptime(
                row["Okuwa giren senesi"].rstrip().lstrip(), "%d.%M.%Y"
            )
        except:
            invalid_fields.append(
                f"Setir №{index + 1}: 'Okuwa giren senesi' meýdançasynda ýalňyşlyk goýberildi"
            )
            row_validation = False
    else:
        data["admission_date"] = row["Okuwa giren senesi"].to_pydatetime()

    return (row_validation, invalid_fields, data)
